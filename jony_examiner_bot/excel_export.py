from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ORANGE = "FFD966"
BLUE_HEADER = "9DC3E6"
YELLOW = "FFFF00"
RED = "FF0000"
GREEN_DARK = "00B050"
GREEN_LIGHT = "92D050"
BLUE_LIGHT = "00B0F0"
GREY = "D9D9D9"

thin = Side(border_style="thin", color="000000")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def status_color(status: str) -> str:
    mapping = {
        "FAIL": RED,
        "BAD": YELLOW,
        "AVERAGE": BLUE_LIGHT,
        "GOOD": GREEN_LIGHT,
        "EXCELLENT": GREEN_DARK,
        "PASS": GREEN_LIGHT,
    }
    return mapping.get(status, "FFFFFF")


def font_color_for_status(status: str):
    """To'q fon (FAIL qizil, EXCELLENT to'q yashil) uchun oq shrift, qolganlari uchun qora."""
    if status in ("FAIL", "EXCELLENT"):
        return "FFFFFF"
    return None


def group_index_color(percent: float) -> str:
    """GROUP INDEX foizini asl baholash shkalasi (FAIL/BAD/AVERAGE/GOOD/EXCELLENT)
    ranglariga moslashtirib bo'yaydi, test turidan qat'i nazar."""
    if percent >= 95:
        return GREEN_DARK
    if percent >= 84:
        return GREEN_LIGHT
    if percent >= 73:
        return BLUE_LIGHT
    if percent >= 64:
        return YELLOW
    return RED


def _cell(ws, row, col, value, bold=False, fill=None, align="center", size=11, font_color=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=size, color=font_color)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = BORDER
    if fill:
        c.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    return c


def build_excel(data: dict, filepath: str):
    """
    data = {
        "teacher": str, "date": str, "level": str,
        "study_dates": str, "examiner": str,
        "test_type": "unit" | "midterm",
        "test_name": str,          # e.g. "UNIT 3" or "END OF COURSE"
        "level_name": str,         # e.g. "NOVA" or "PRIME"
        "max_score": int,          # unit only
        "sections": {"listening":20,"reading":25,"writing":40,"speaking":15},  # midterm only
        "students": [ {surname, name, ..., total, percent, status, first_time} ]
    }
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Natijalar"

    is_unit = data["test_type"] == "unit"
    n_cols = 6 if is_unit else 9
    last_col_letter = get_column_letter(n_cols)

    # Title
    ws.merge_cells(f"A1:{last_col_letter}1")
    _cell(ws, 1, 1, "JONY ACADEMY - EXAM RESULTS", bold=True, size=16)
    ws.row_dimensions[1].height = 28

    header_rows = [
        ("TEACHER", data["teacher"]),
        ("DATE", data["date"]),
        ("LEVEL", data["level"]),
        ("NUMBER OF STUDENTS", str(len(data["students"]))),
        ("STUDY DATES AND TIMES", data["study_dates"]),
        ("EXAMINER", data["examiner"]),
    ]

    r = 2
    for label, val in header_rows:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        _cell(ws, r, 1, label, bold=True, fill=ORANGE)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=n_cols)
        _cell(ws, r, 3, val, bold=True)
        r += 1

    # Test type / section header row
    if is_unit:
        _cell(ws, r, 1, data["test_name"], bold=True, fill=ORANGE)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        _cell(ws, r, 2, data["level_name"], bold=True)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=n_cols)
        _cell(ws, r, 4, str(data["max_score"]), bold=True)
        r += 1

        headers = ["#", "SURNAME", "NAME", "TOTAL", "PERCENT", "STATUS"]
    else:
        sec = data["sections"]
        _cell(ws, r, 1, data["test_name"], bold=True, fill=ORANGE)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        _cell(ws, r, 2, data["level_name"], bold=True)
        _cell(ws, r, 4, sec["listening"], bold=True)
        _cell(ws, r, 5, sec["reading"], bold=True)
        _cell(ws, r, 6, sec["writing"], bold=True)
        _cell(ws, r, 7, sec["speaking"], bold=True)
        total_max = sum(sec.values())
        ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=9)
        _cell(ws, r, 8, total_max, bold=True)
        r += 1

        headers = ["#", "SURNAME", "NAME", "LISTENING", "READING", "WRITING", "SPEAKING", "TOTAL", "STATUS"]

    for i, h in enumerate(headers, start=1):
        _cell(ws, r, i, h, bold=True, fill=BLUE_HEADER)
    header_row = r
    r += 1

    all_students = data["students"]
    # Checkbox orqali "QAYTA TOPSHIRMOQDA" deb belgilanganlar (first_time=False)
    # asosiy jadvalda + GROUP/PASSING INDEKSGA kiradi.
    index_students = sorted(
        [s for s in all_students if not s["first_time"]],
        key=lambda s: s["percent"], reverse=True,
    )
    # Belgilanmaganlar (first_time=True) — bitta qator tashlab, pastda ALOHIDA
    # jadvalda ko'rsatiladi, indeksga kirmaydi.
    separate_students = sorted(
        [s for s in all_students if s["first_time"]],
        key=lambda s: s["percent"], reverse=True,
    )

    def write_student(row_idx, idx, s):
        fc = font_color_for_status(s["status"])
        if is_unit:
            _cell(ws, row_idx, 1, idx, font_color=fc)
            _cell(ws, row_idx, 2, s["surname"], align="left", font_color=fc)
            _cell(ws, row_idx, 3, s["name"], align="left", font_color=fc)
            _cell(ws, row_idx, 4, s["total"], font_color=fc)
            _cell(ws, row_idx, 5, f'{s["percent"]:.1f}%', font_color=fc)
            _cell(ws, row_idx, 6, s["status"], bold=True, font_color=fc)
        else:
            _cell(ws, row_idx, 1, idx, font_color=fc)
            _cell(ws, row_idx, 2, s["surname"], align="left", font_color=fc)
            _cell(ws, row_idx, 3, s["name"], align="left", font_color=fc)
            _cell(ws, row_idx, 4, s["listening"], font_color=fc)
            _cell(ws, row_idx, 5, s["reading"], font_color=fc)
            _cell(ws, row_idx, 6, s["writing"], font_color=fc)
            _cell(ws, row_idx, 7, s["speaking"], font_color=fc)
            _cell(ws, row_idx, 8, s["total"], font_color=fc)
            _cell(ws, row_idx, 9, s["status"], bold=True, font_color=fc)
        # Butun qatorni status rangiga bo'yash
        color = status_color(s["status"])
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, n_cols + 1):
            ws.cell(row=row_idx, column=c).fill = fill

    for idx, s in enumerate(index_students, start=1):
        write_student(r, idx, s)
        r += 1

    # Alohida jadval: BIRINCHI MARTA topshirganlar (checkbox belgilanmagan)
    # — asosiy jadvaldan bitta bo'sh qator tashlab, GROUP/PASSING INDEXdan OLDIN
    # ko'rsatiladi, lekin indeksga kirmaydi
    if separate_students:
        r += 1  # bitta bo'sh qator
        for idx, s in enumerate(separate_students, start=1):
            write_student(r, idx, s)
            r += 1

    # Group / Passing index — faqat checkbox bilan belgilangan (first_time=False) asosida
    r += 1
    if index_students:
        avg_percent = sum(s["percent"] for s in index_students) / len(index_students)
        passed = [s for s in index_students if s["status"] not in ("FAIL",)]
        passing_percent = len(passed) / len(index_students) * 100
    else:
        avg_percent = 0
        passing_percent = 0

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols - 2)
    _cell(ws, r, 1, "GROUP INDEX", bold=True, fill=YELLOW, align="right")
    ws.merge_cells(start_row=r, start_column=n_cols - 1, end_row=r, end_column=n_cols)
    _cell(ws, r, n_cols - 1, f"{avg_percent:.0f}%", bold=True, fill=group_index_color(avg_percent))
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols - 2)
    _cell(ws, r, 1, "PASSING INDEX", bold=True, fill=YELLOW, align="right")
    ws.merge_cells(start_row=r, start_column=n_cols - 1, end_row=r, end_column=n_cols)
    _cell(ws, r, n_cols - 1, f"{passing_percent:.1f}%", bold=True, fill=YELLOW)
    r += 1

    # Column widths
    widths = [4, 18, 16, 11, 11, 11, 11, 9, 11] if not is_unit else [4, 18, 16, 9, 10, 11]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(filepath)
    return filepath
